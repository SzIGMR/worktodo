"""Shared IO helpers for skillrunner."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import load_workbook


NORMALIZE_PATTERN = re.compile(r"[^a-z0-9]+")


def normalize_column(name: str) -> str:
    """Normalize a column name for matching."""
    name = name.strip().lower()
    name = NORMALIZE_PATTERN.sub("_", name)
    return name.strip("_")


def parse_german_number(value: str) -> float:
    """Parse German formatted numbers like ' 7.452,00 € '."""
    cleaned = value.strip().replace("\u00a0", " ")
    if cleaned == "":
        raise ValueError("Leerer Betrag")
    cleaned = re.sub(r"(?i)\b(?:eur|euro)\b", "", cleaned)
    cleaned = cleaned.replace("€", "")
    cleaned = cleaned.replace(" ", "")
    cleaned = re.sub(r"[^\d,.\-+]", "", cleaned)
    if cleaned == "" or cleaned in {"+", "-"}:
        raise ValueError("Leerer Betrag")
    if "," in cleaned:
        cleaned = cleaned.replace(".", "")
        cleaned = cleaned.replace(",", ".")
    elif "." in cleaned:
        if cleaned.count(".") > 1:
            cleaned = cleaned.replace(".", "")
        else:
            left, right = cleaned.split(".")
            if len(right) == 3 and left:
                cleaned = f"{left}{right}"
    return float(cleaned)


def sniff_csv_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample)
        return dialect.delimiter
    except csv.Error:
        return ","


@dataclass
class TableData:
    headers: list[str]
    rows: list[list[str]]


def read_csv_table(path: Path) -> TableData:
    content = path.read_text(encoding="utf-8")
    delimiter = sniff_csv_delimiter(content[:4096])
    reader = csv.reader(content.splitlines(), delimiter=delimiter)
    rows = list(reader)
    if not rows:
        raise ValueError("CSV-Datei ist leer")
    headers = rows[0]
    return TableData(headers=headers, rows=rows[1:])


def read_excel_table(path: Path, sheet_name: str | None = None) -> TableData:
    if path.suffix.lower() == ".xls":
        return _read_xls_table(path, sheet_name)
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel-Datei ist leer")
    headers = [str(cell or "") for cell in rows[0]]
    data_rows = []
    for row in rows[1:]:
        data_rows.append(["" if cell is None else str(cell) for cell in row])
    return TableData(headers=headers, rows=data_rows)


def _read_xls_table(path: Path, sheet_name: str | None = None) -> TableData:
    try:
        import xlrd
    except ModuleNotFoundError as exc:  # pragma: no cover - optional dependency
        raise ValueError("Für .xls Dateien bitte 'xlrd' installieren.") from exc
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)
    if sheet.nrows == 0:
        raise ValueError("Excel-Datei ist leer")
    headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
    data_rows = []
    for row_idx in range(1, sheet.nrows):
        data_rows.append([str(sheet.cell_value(row_idx, col)) for col in range(sheet.ncols)])
    return TableData(headers=headers, rows=data_rows)


def read_table(path: Path) -> TableData:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_table(path)
    if suffix in {".xlsx", ".xls"}:
        return read_excel_table(path)
    raise ValueError("Unbekanntes Dateiformat. Unterstützt: .csv, .xlsx, .xls")


def map_columns(headers: Sequence[str]) -> dict[str, str]:
    return {normalize_column(h): h for h in headers}


def rows_as_dicts(table: TableData) -> list[dict[str, str]]:
    return [dict(zip(table.headers, row, strict=False)) for row in table.rows]


def select_columns(
    table: TableData, required: Iterable[str], mapping: dict[str, str]
) -> TableData:
    missing = [col for col in required if col not in mapping]
    if missing:
        raise ValueError(
            "Fehlende Spalten: " + ", ".join(missing)
        )
    headers = [mapping[col] for col in required]
    index = [table.headers.index(header) for header in headers]
    rows = []
    for row in table.rows:
        rows.append([row[i] if i < len(row) else "" for i in index])
    return TableData(headers=headers, rows=rows)
