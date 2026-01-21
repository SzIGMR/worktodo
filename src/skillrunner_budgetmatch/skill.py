"""Budget matcher skill implementation."""

from __future__ import annotations

import argparse
import math
import statistics
import time
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook

from skillrunner.excel_utils import write_sheet
from skillrunner.io_utils import (
    TableData,
    map_columns,
    normalize_column,
    parse_german_number,
    read_table,
)


@dataclass(frozen=True)
class Booking:
    person: str
    month: date
    amount_eur: float


@dataclass(frozen=True)
class Budget:
    project: str
    budget_eur: float


@dataclass
class AllocationResult:
    allocation_rows: list[dict[str, object]]
    project_summary: list[dict[str, object]]
    monthly_summary: list[dict[str, object]]
    unallocated: list[dict[str, object]]
    diagnostics: dict[str, object]


class BudgetMatcherSkill:
    name = "budget_matcher"
    description = "Abgleich von Buchungen mit Projektbudgets"

    def cli_help(self) -> str:
        return "Budgetabgleich und Forecasting für Projektbudgets"

    def run(self, args: list[str]) -> int:
        parser = build_parser()
        opts = parser.parse_args(args)

        if opts.generate_templates:
            generate_templates(Path(opts.generate_templates))
            print("Vorlagen wurden erstellt.")
            return 0

        if opts.convert_inputs:
            if not opts.bookings or not opts.budgets:
                print("Für die Konvertierung sind --bookings und --budgets erforderlich.")
                return 2
            convert_inputs(Path(opts.bookings), Path(opts.budgets), Path(opts.convert_inputs))
            print("Eingaben wurden konvertiert.")
            return 0

        if not opts.bookings or not opts.budgets or not opts.out:
            print("Bitte --bookings, --budgets und --out angeben.")
            return 2

        bookings_table = read_table(Path(opts.bookings))
        budgets_table = read_table(Path(opts.budgets))

        bookings, bookings_rows, out_of_period = parse_bookings(
            bookings_table,
            fy_start_year=opts.fy_start_year,
            include_out_of_period=opts.include_out_of_period,
        )
        budgets, budgets_rows, warnings = parse_budgets(
            budgets_table,
            min_project_rest=opts.min_project_rest_eur,
        )

        if opts.mode == "optimize":
            result = run_optimization(
                bookings=bookings,
                budgets=budgets,
                fy_start_year=opts.fy_start_year,
                min_project_rest=opts.min_project_rest_eur,
                person_max_projects=opts.person_max_projects,
                person_project_penalty=opts.person_project_penalty,
                monthly_smoothness_penalty=opts.monthly_smoothness_penalty,
                unallocated_person_penalty=opts.unallocated_person_penalty,
                solver=opts.solver,
                time_limit_seconds=opts.time_limit_seconds,
            )
        else:
            result = run_forecast(
                bookings=bookings,
                budgets=budgets,
                fy_start_year=opts.fy_start_year,
            )

        diagnostics = {**result.diagnostics, "warnings": warnings}
        write_workbook(
            Path(opts.out),
            budgets_rows=budgets_rows,
            bookings_rows=bookings_rows,
            allocation=result.allocation_rows,
            project_summary=result.project_summary,
            monthly_summary=result.monthly_summary,
            unallocated=result.unallocated,
            diagnostics=diagnostics,
            out_of_period=out_of_period,
            forecast_rows=result.diagnostics.get("forecast_rows"),
        )
        return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="skillrunner run budget_matcher",
        description=(
            "Allokiert Buchungen auf Projektbudgets oder erstellt eine Forecast-Analyse."
        ),
    )
    parser.add_argument("--bookings", help="Pfad zur Buchungsdatei (.csv/.xlsx)")
    parser.add_argument("--budgets", help="Pfad zur Budgetdatei (.csv/.xlsx)")
    parser.add_argument("--out", help="Ausgabe-Excel-Datei")
    parser.add_argument("--mode", choices=["optimize", "forecast"], default="optimize")
    parser.add_argument("--fy-start-year", type=int, required=True)
    parser.add_argument("--min-project-rest-eur", type=float, default=1.0)
    parser.add_argument("--person-max-projects", type=int, default=2)
    parser.add_argument("--person-project-penalty", type=float, default=10.0)
    parser.add_argument("--monthly-smoothness-penalty", type=float, default=0.1)
    parser.add_argument("--unallocated-person-penalty", type=float, default=0.01)
    parser.add_argument("--include-out-of-period", type=str, default="false")
    parser.add_argument("--solver", choices=["auto", "pulp", "ortools", "heuristic"], default="auto")
    parser.add_argument("--time-limit-seconds", type=int, default=30)
    parser.add_argument("--write-intermediate-csv", type=str, default="false")
    parser.add_argument("--generate-templates", help="Zielordner für Vorlagen")
    parser.add_argument("--convert-inputs", help="Zielordner für konvertierte Eingaben")
    return parser


def parse_bool(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "ja"}


def parse_month(value: str) -> date:
    if "/" not in value:
        raise ValueError(f"Ungültiges Monatsformat: {value}")
    year_str, month_str = value.split("/")
    return date(int(year_str), int(month_str), 1)


def parse_bookings(
    table: TableData,
    fy_start_year: int,
    include_out_of_period: str | bool,
) -> tuple[list[Booking], list[dict[str, object]], list[dict[str, object]]]:
    include_flag = parse_bool(str(include_out_of_period))
    normalized = map_columns(table.headers)
    required = {
        "vorname": ["vorname", "name", "person"],
        "bezugsmonat": ["bezugsmonat", "monat", "month"],
        "eur": ["eur", "betrag", "amount"],
    }
    mapped = _map_required_columns(normalized, required)

    bookings: list[Booking] = []
    rows: list[dict[str, object]] = []
    out_of_period: list[dict[str, object]] = []
    months = fiscal_year_months(fy_start_year)

    for row in table.rows:
        row_map = dict(zip(table.headers, row, strict=False))
        person = row_map.get(mapped["vorname"], "").strip()
        month_raw = row_map.get(mapped["bezugsmonat"], "").strip()
        amount_raw = row_map.get(mapped["eur"], "").strip()
        month = parse_month(month_raw)
        amount = parse_german_number(amount_raw)
        is_out = month not in months
        record = {
            "person": person,
            "month": month,
            "amount_eur": amount,
            "out_of_period": is_out,
        }
        if is_out and not include_flag:
            out_of_period.append(record)
            continue
        bookings.append(Booking(person=person, month=month, amount_eur=amount))
        rows.append(record)

    return bookings, rows, out_of_period


def parse_budgets(
    table: TableData,
    min_project_rest: float,
) -> tuple[list[Budget], list[dict[str, object]], list[str]]:
    normalized = map_columns(table.headers)
    required = {
        "projekt": ["projekt", "project"],
        "bewilligt": ["bewilligt", "budget"],
    }
    mapped = _map_required_columns(normalized, required)
    budgets: list[Budget] = []
    rows: list[dict[str, object]] = []
    warnings: list[str] = []

    for row in table.rows:
        row_map = dict(zip(table.headers, row, strict=False))
        project = row_map.get(mapped["projekt"], "").strip()
        amount_raw = row_map.get(mapped["bewilligt"], "").strip()
        amount = parse_german_number(amount_raw)
        if amount <= min_project_rest:
            warnings.append(
                f"Projekt {project} ist aufgrund des Mindestrests nicht nutzbar."
            )
        budgets.append(Budget(project=project, budget_eur=amount))
        rows.append({"project": project, "budget_eur": amount})

    return budgets, rows, warnings


def _map_required_columns(
    normalized: dict[str, str], required: dict[str, list[str]]
) -> dict[str, str]:
    mapped: dict[str, str] = {}
    for key, candidates in required.items():
        found = None
        for candidate in candidates:
            norm = normalize_column(candidate)
            if norm in normalized:
                found = normalized[norm]
                break
        if not found:
            raise ValueError(f"Fehlende Spalte: {key}")
        mapped[key] = found
    return mapped


def fiscal_year_months(fy_start_year: int) -> list[date]:
    months = []
    for i in range(12):
        month = ((4 + i - 1) % 12) + 1
        year = fy_start_year if month >= 4 else fy_start_year + 1
        months.append(date(year, month, 1))
    return months


def run_optimization(
    bookings: list[Booking],
    budgets: list[Budget],
    fy_start_year: int,
    min_project_rest: float,
    person_max_projects: int,
    person_project_penalty: float,
    monthly_smoothness_penalty: float,
    unallocated_person_penalty: float,
    solver: str,
    time_limit_seconds: int,
) -> AllocationResult:
    start = time.time()
    if solver in {"auto", "pulp", "ortools"}:
        try:
            return _solve_with_pulp(
                bookings,
                budgets,
                fy_start_year,
                min_project_rest,
                person_max_projects,
                person_project_penalty,
                monthly_smoothness_penalty,
                unallocated_person_penalty,
                solver,
                time_limit_seconds,
                start,
            )
        except ModuleNotFoundError:
            pass

    return _solve_heuristic(
        bookings,
        budgets,
        fy_start_year,
        min_project_rest,
        person_max_projects,
        person_project_penalty,
        monthly_smoothness_penalty,
        unallocated_person_penalty,
        start,
    )


def _solve_with_pulp(
    bookings: list[Booking],
    budgets: list[Budget],
    fy_start_year: int,
    min_project_rest: float,
    person_max_projects: int,
    person_project_penalty: float,
    monthly_smoothness_penalty: float,
    unallocated_person_penalty: float,
    solver: str,
    time_limit_seconds: int,
    start: float,
) -> AllocationResult:
    import pulp

    problem = pulp.LpProblem("budget_matcher", pulp.LpMaximize)

    projects = [b for b in budgets if b.budget_eur > min_project_rest]
    project_names = [p.project for p in projects]
    booking_indices = list(range(len(bookings)))

    x = pulp.LpVariable.dicts("assign", (booking_indices, project_names), cat="Binary")

    for i in booking_indices:
        problem += (
            pulp.lpSum(x[i][j] for j in project_names) <= 1,
            f"assign_{i}",
        )

    for project in projects:
        problem += (
            pulp.lpSum(
                bookings[i].amount_eur * x[i][project.project]
                for i in booking_indices
            )
            <= project.budget_eur - min_project_rest,
            f"cap_{project.project}",
        )

    persons = sorted({b.person for b in bookings})
    y = pulp.LpVariable.dicts("person_project", (persons, project_names), cat="Binary")
    for p in persons:
        indices = [i for i, b in enumerate(bookings) if b.person == p]
        for proj in project_names:
            problem += (
                pulp.lpSum(x[i][proj] for i in indices)
                <= len(indices) * y[p][proj],
                f"link_{p}_{proj}",
            )

    excess = pulp.LpVariable.dicts("excess_projects", persons, lowBound=0)
    for p in persons:
        problem += (
            excess[p]
            >= pulp.lpSum(y[p][proj] for proj in project_names) - person_max_projects,
            f"excess_{p}",
        )

    months = fiscal_year_months(fy_start_year)
    month_set = set(months)
    m = pulp.LpVariable.dicts("monthly", (project_names, months), lowBound=0)
    for proj in project_names:
        for month in months:
            indices = [i for i, b in enumerate(bookings) if b.month == month]
            problem += (
                m[proj][month]
                == pulp.lpSum(bookings[i].amount_eur * x[i][proj] for i in indices),
                f"monthly_{proj}_{month}",
            )

    diff = pulp.LpVariable.dicts("smooth_diff", (project_names, months[1:]), lowBound=0)
    for proj in project_names:
        for month in months[1:]:
            prev_month = months[months.index(month) - 1]
            problem += diff[proj][month] >= m[proj][month] - m[proj][prev_month]
            problem += diff[proj][month] >= m[proj][prev_month] - m[proj][month]

    u = pulp.LpVariable.dicts("unalloc_person", persons, cat="Binary")
    for p in persons:
        indices = [i for i, b in enumerate(bookings) if b.person == p]
        for i in indices:
            problem += 1 - pulp.lpSum(x[i][j] for j in project_names) <= u[p]

    total_allocated = pulp.lpSum(
        bookings[i].amount_eur * x[i][j] for i in booking_indices for j in project_names
    )
    penalty_person = person_project_penalty * pulp.lpSum(excess[p] for p in persons)
    penalty_smooth = monthly_smoothness_penalty * pulp.lpSum(
        diff[proj][month] for proj in project_names for month in months[1:]
    )
    penalty_unalloc = unallocated_person_penalty * pulp.lpSum(u[p] for p in persons)

    problem += 1_000_000 * total_allocated - 1_000 * penalty_person - penalty_smooth - 0.1 * penalty_unalloc

    if solver in {"auto", "pulp"}:
        pulp_solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=time_limit_seconds)
    else:
        pulp_solver = pulp.PULP_CBC_CMD(msg=False, timeLimit=time_limit_seconds)

    status = problem.solve(pulp_solver)
    runtime = time.time() - start

    allocation_rows = []
    for i, booking in enumerate(bookings):
        assigned = "UNALLOCATED"
        for proj in project_names:
            if pulp.value(x[i][proj]) > 0.5:
                assigned = proj
                break
        allocation_rows.append(
            {
                "person": booking.person,
                "month": booking.month,
                "amount_eur": booking.amount_eur,
                "assigned_project": assigned,
                "status": "allocated" if assigned != "UNALLOCATED" else "unallocated",
                "notes": "out_of_period" if booking.month not in month_set else "",
            }
        )

    result = _build_summaries(bookings, budgets, allocation_rows, fy_start_year, min_project_rest)
    result.diagnostics.update(
        {
            "solver_used": "pulp",
            "status": pulp.LpStatus[status],
            "runtime_seconds": runtime,
            "objective": pulp.value(problem.objective),
            "total_allocated": pulp.value(total_allocated),
            "penalty_person_projects": pulp.value(penalty_person),
            "penalty_smooth": pulp.value(penalty_smooth),
            "penalty_unallocated_persons": pulp.value(penalty_unalloc),
            "weights": {
                "allocated": 1_000_000,
                "person": 1_000,
                "smooth": 1,
                "unallocated": 0.1,
            },
        }
    )
    return result


def _solve_heuristic(
    bookings: list[Booking],
    budgets: list[Budget],
    fy_start_year: int,
    min_project_rest: float,
    person_max_projects: int,
    person_project_penalty: float,
    monthly_smoothness_penalty: float,
    unallocated_person_penalty: float,
    start: float,
) -> AllocationResult:
    remaining = {b.project: b.budget_eur - min_project_rest for b in budgets}
    allocation_rows: list[dict[str, object]] = []
    bookings_sorted = sorted(bookings, key=lambda b: b.amount_eur, reverse=True)
    month_set = set(fiscal_year_months(fy_start_year))

    for booking in bookings_sorted:
        best_project = None
        for project, rest in sorted(remaining.items(), key=lambda item: item[1], reverse=True):
            if rest >= booking.amount_eur:
                best_project = project
                break
        if best_project:
            remaining[best_project] -= booking.amount_eur
            assigned = best_project
        else:
            assigned = "UNALLOCATED"
        allocation_rows.append(
            {
                "person": booking.person,
                "month": booking.month,
                "amount_eur": booking.amount_eur,
                "assigned_project": assigned,
                "status": "allocated" if assigned != "UNALLOCATED" else "unallocated",
                "notes": "out_of_period" if booking.month not in month_set else "Heuristik",
            }
        )

    result = _build_summaries(bookings, budgets, allocation_rows, fy_start_year, min_project_rest)
    runtime = time.time() - start
    result.diagnostics.update(
        {
            "solver_used": "heuristic",
            "status": "heuristic",
            "runtime_seconds": runtime,
            "objective": None,
            "total_allocated": sum(
                row["amount_eur"] for row in allocation_rows if row["assigned_project"] != "UNALLOCATED"
            ),
            "penalty_person_projects": None,
            "penalty_smooth": None,
            "penalty_unallocated_persons": None,
            "weights": {
                "allocated": 1_000_000,
                "person": 1_000,
                "smooth": 1,
                "unallocated": 0.1,
            },
        }
    )
    return result


def _build_summaries(
    bookings: list[Booking],
    budgets: list[Budget],
    allocation_rows: list[dict[str, object]],
    fy_start_year: int,
    min_project_rest: float,
) -> AllocationResult:
    allocated_by_project = {b.project: 0.0 for b in budgets}
    for row in allocation_rows:
        project = row["assigned_project"]
        if project != "UNALLOCATED":
            allocated_by_project[project] += float(row["amount_eur"])

    project_summary = []
    for budget in budgets:
        allocated = allocated_by_project.get(budget.project, 0.0)
        remaining = budget.budget_eur - allocated
        utilization = (allocated / budget.budget_eur * 100) if budget.budget_eur else 0.0
        project_summary.append(
            {
                "project": budget.project,
                "budget_eur": budget.budget_eur,
                "allocated_eur": allocated,
                "remaining_eur": remaining,
                "min_rest_eur": min_project_rest,
                "utilization_pct": utilization,
            }
        )

    months = fiscal_year_months(fy_start_year)
    monthly_summary = []
    for month in months:
        row = {"month": month}
        for budget in budgets:
            row[budget.project] = 0.0
        row["UNALLOCATED"] = 0.0
        monthly_summary.append(row)

    month_lookup = {row["month"]: row for row in monthly_summary}
    for row in allocation_rows:
        target = row["assigned_project"]
        month = row["month"]
        if month in month_lookup:
            month_lookup[month][target] = month_lookup[month].get(target, 0.0) + float(
                row["amount_eur"]
            )

    unallocated = [row for row in allocation_rows if row["assigned_project"] == "UNALLOCATED"]

    return AllocationResult(
        allocation_rows=allocation_rows,
        project_summary=project_summary,
        monthly_summary=monthly_summary,
        unallocated=unallocated,
        diagnostics={},
    )


def run_forecast(
    bookings: list[Booking],
    budgets: list[Budget],
    fy_start_year: int,
) -> AllocationResult:
    months = fiscal_year_months(fy_start_year)
    monthly_totals = {month: 0.0 for month in months}
    for booking in bookings:
        if booking.month in monthly_totals:
            monthly_totals[booking.month] += booking.amount_eur

    month_values = [monthly_totals[month] for month in months if monthly_totals[month] > 0]
    actuals_to_date = sum(month_values)
    months_with_data = len(month_values)
    total_months = len(months)

    if months_with_data >= 3:
        indices = list(range(months_with_data))
        mean_x = statistics.mean(indices)
        mean_y = statistics.mean(month_values)
        numerator = sum((x - mean_x) * (y - mean_y) for x, y in zip(indices, month_values))
        denominator = sum((x - mean_x) ** 2 for x in indices) or 1
        slope = numerator / denominator
        intercept = mean_y - slope * mean_x
        trend_forecast = sum(
            max(0.0, intercept + slope * i) for i in range(total_months)
        )
    else:
        trend_forecast = 0.0

    run_rate = (actuals_to_date / months_with_data * total_months) if months_with_data else 0.0
    forecast_total = trend_forecast if months_with_data >= 3 else run_rate
    confidence_low = run_rate
    confidence_high = forecast_total * 1.05 if forecast_total else run_rate
    total_budget = sum(b.budget_eur for b in budgets)
    delta_vs_budget = forecast_total - total_budget

    allocation_rows = [
        {
            "person": b.person,
            "month": b.month,
            "amount_eur": b.amount_eur,
            "assigned_project": "UNALLOCATED",
            "status": "forecast",
            "notes": "",
        }
        for b in bookings
    ]

    result = _build_summaries(bookings, budgets, allocation_rows, fy_start_year, 0.0)
    result.diagnostics.update(
        {
            "solver_used": "forecast",
            "status": "forecast",
            "forecast_rows": [
                {
                    "actuals_to_date": actuals_to_date,
                    "forecast_total": forecast_total,
                    "confidence_low": confidence_low,
                    "confidence_high": confidence_high,
                    "delta_vs_budget": delta_vs_budget,
                }
            ],
            "forecast_methods": {
                "trend": trend_forecast,
                "run_rate": run_rate,
                "months_with_data": months_with_data,
            },
        }
    )
    return result


def generate_templates(folder: Path) -> None:
    folder.mkdir(parents=True, exist_ok=True)
    _write_template(
        folder / "bookings_template.xlsx",
        headers=["Vorname", "Bezugsmonat", "EUR"],
        example=["Max Mustermann", "2025/10", "1.234,00"],
        instructions="Spalten: Vorname, Bezugsmonat (YYYY/MM), EUR (deutsches Zahlenformat).",
    )
    _write_template(
        folder / "budgets_template.xlsx",
        headers=["Projekt", "Bewilligt"],
        example=["Projekt A", "7.452,00 €"],
        instructions="Spalten: Projekt, Bewilligt (deutsches Zahlenformat).",
    )


def _write_template(path: Path, headers: list[str], example: list[str], instructions: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(headers)
    ws.append(example)
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.append([instructions])
    wb.save(path)


def convert_inputs(bookings_path: Path, budgets_path: Path, out_folder: Path) -> None:
    out_folder.mkdir(parents=True, exist_ok=True)
    bookings_table = read_table(bookings_path)
    budgets_table = read_table(budgets_path)

    converted_bookings = _convert_table(
        bookings_table,
        {
            "Vorname": ["vorname", "name", "person"],
            "Bezugsmonat": ["bezugsmonat", "monat", "month"],
            "EUR": ["eur", "betrag", "amount"],
        },
    )
    converted_budgets = _convert_table(
        budgets_table,
        {
            "Projekt": ["projekt", "project"],
            "Bewilligt": ["bewilligt", "budget"],
        },
    )

    _write_converted(out_folder / "bookings_converted.xlsx", converted_bookings)
    _write_converted(out_folder / "budgets_converted.xlsx", converted_budgets)


def _convert_table(table: TableData, mapping: dict[str, list[str]]) -> TableData:
    normalized = map_columns(table.headers)
    mapped = {}
    for target, candidates in mapping.items():
        for candidate in candidates:
            norm = normalize_column(candidate)
            if norm in normalized:
                mapped[target] = normalized[norm]
                break
        if target not in mapped:
            mapped[target] = table.headers[0]
    rows = []
    for row in table.rows:
        row_map = dict(zip(table.headers, row, strict=False))
        rows.append([row_map.get(mapped[target], "") for target in mapping.keys()])
    return TableData(headers=list(mapping.keys()), rows=rows)


def _write_converted(path: Path, table: TableData) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Converted"
    ws.append(table.headers)
    for row in table.rows:
        ws.append(row)
    wb.save(path)


def write_workbook(
    path: Path,
    budgets_rows: list[dict[str, object]],
    bookings_rows: list[dict[str, object]],
    allocation: list[dict[str, object]],
    project_summary: list[dict[str, object]],
    monthly_summary: list[dict[str, object]],
    unallocated: list[dict[str, object]],
    diagnostics: dict[str, object],
    out_of_period: list[dict[str, object]],
    forecast_rows: list[dict[str, object]] | None,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    write_sheet(
        wb,
        "Inputs_Budgets",
        headers=["project", "budget_eur"],
        rows=[[row["project"], row["budget_eur"]] for row in budgets_rows],
    )
    write_sheet(
        wb,
        "Inputs_Bookings",
        headers=["person", "month", "amount_eur", "out_of_period"],
        rows=[
            [row["person"], row["month"], row["amount_eur"], row["out_of_period"]]
            for row in bookings_rows
        ],
    )
    write_sheet(
        wb,
        "Allocation",
        headers=["person", "month", "amount_eur", "assigned_project", "status", "notes"],
        rows=[
            [
                row["person"],
                row["month"],
                row["amount_eur"],
                row["assigned_project"],
                row["status"],
                row["notes"],
            ]
            for row in allocation
        ],
    )
    write_sheet(
        wb,
        "Project_Summary",
        headers=[
            "project",
            "budget_eur",
            "allocated_eur",
            "remaining_eur",
            "min_rest_eur",
            "utilization_pct",
        ],
        rows=[
            [
                row["project"],
                row["budget_eur"],
                row["allocated_eur"],
                row["remaining_eur"],
                row["min_rest_eur"],
                row["utilization_pct"],
            ]
            for row in project_summary
        ],
    )
    write_sheet(
        wb,
        "Unallocated",
        headers=["person", "month", "amount_eur", "assigned_project", "status", "notes"],
        rows=[
            [
                row["person"],
                row["month"],
                row["amount_eur"],
                row["assigned_project"],
                row["status"],
                row["notes"],
            ]
            for row in unallocated
        ],
    )

    monthly_headers = ["month"] + [row["project"] for row in project_summary] + ["UNALLOCATED"]
    monthly_rows = []
    for row in monthly_summary:
        monthly_rows.append([row.get(header, 0.0) for header in monthly_headers])
    write_sheet(
        wb,
        "Monthly_Project_Summary",
        headers=monthly_headers,
        rows=monthly_rows,
    )

    write_sheet(
        wb,
        "Diagnostics",
        headers=["key", "value"],
        rows=[[key, value] for key, value in diagnostics.items()],
    )

    if out_of_period:
        write_sheet(
            wb,
            "Out_of_Period",
            headers=["person", "month", "amount_eur"],
            rows=[[row["person"], row["month"], row["amount_eur"]] for row in out_of_period],
        )

    if forecast_rows:
        write_sheet(
            wb,
            "Forecast",
            headers=[
                "actuals_to_date",
                "forecast_total",
                "confidence_low",
                "confidence_high",
                "delta_vs_budget",
            ],
            rows=[
                [
                    row["actuals_to_date"],
                    row["forecast_total"],
                    row["confidence_low"],
                    row["confidence_high"],
                    row["delta_vs_budget"],
                ]
                for row in forecast_rows
            ],
        )

    wb.save(path)
